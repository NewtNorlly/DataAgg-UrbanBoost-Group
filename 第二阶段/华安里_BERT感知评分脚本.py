"""
华安里感知模型 - BERT情感分类与6+1框架加权计算
任务2：感知模型实现与分数计算
"""

# !! 必须在所有import之前设置HF镜像，否则transformers会用默认URL
import os
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

import json, sys, time
import pandas as pd
import openpyxl
from transformers import pipeline
import torch

# ============ 配置 ============
XLSX_PATH = r"C:\Users\NewtN\下载\华安里_分析结果.xlsx"
OUTPUT_CSV = r"C:\Honey\华安里_感知得分.csv"
OUTPUT_EXCEL = r"C:\Honey\华安里_感知得分.xlsx"
OUTPUT_JSON = r"C:\Honey\华安里_感知得分.json"

# ============ 维度映射：AI 30指标 → 6+1框架7维度 ============
# 框架维度权重（来自任务1同学交付的贝叶斯校准后验均值）
DIMENSION_WEIGHTS = {
    "街巷空间与交通通达": 0.16,
    "建筑风貌与界面品质": 0.13,
    "便民生活与服务韧性": 0.15,
    "环境秩序与生态韧性": 0.14,
    "公共安全与心理感知": 0.13,
    "场所氛围与情感价值": 0.14,
    "针灸式更新潜力": 0.15,
}

# AI维度 → 框架维度 的映射
# 同时标记哪些是负向指标（指标值越高代表品质越差，需要在计算时反转）
AI_TO_FRAMEWORK = {
    # 维度1：街巷空间与交通通达
    "交通可达性": ("街巷空间与交通通达", False),
    "交通秩序度": ("街巷空间与交通通达", False),
    "休憩与交往空间": ("街巷空间与交通通达", False),
    "无障碍设施": ("街巷空间与交通通达", False),
    "空间开阔度": ("街巷空间与交通通达", False),
    
    # 维度2：建筑风貌与界面品质
    "视觉多样性": ("建筑风貌与界面品质", False),
    "美感评价": ("建筑风貌与界面品质", False),
    "空间整洁度": ("建筑风貌与界面品质", False),  # 整洁度=界面品质
    
    # 维度3：便民生活与服务韧性
    "商业多样性": ("便民生活与服务韧性", False),
    "街道繁荣度": ("便民生活与服务韧性", False),
    "摊贩活动": ("便民生活与服务韧性", False),
    "商业环境": ("便民生活与服务韧性", False),
    "经济活力": ("便民生活与服务韧性", False),
    "社会与空间公平": ("便民生活与服务韧性", False),
    
    # 维度4：环境秩序与生态韧性
    "绿化覆盖率": ("环境秩序与生态韧性", False),
    "植物健康度": ("环境秩序与生态韧性", False),
    "水体可见度": ("环境秩序与生态韧性", False),
    "生态廊道连通性": ("环境秩序与生态韧性", False),
    "海绵城市元素": ("环境秩序与生态韧性", False),
    "污染可见度": ("环境秩序与生态韧性", True),   # 负向：越多污染越差
    "热岛效应迹象": ("环境秩序与生态韧性", True),  # 负向：越明显越差
    
    # 维度5：公共安全与心理感知
    "噪音干扰度": ("公共安全与心理感知", True),     # 负向：越多噪音越差
    "负面情绪度": ("公共安全与心理感知", True),     # 负向：越高越差
    
    # 维度6：场所氛围与情感价值
    "景观特色度": ("场所氛围与情感价值", False),
    "自然舒适感": ("场所氛围与情感价值", False),
    "文化符号": ("场所氛围与情感价值", False),
    "历史文脉": ("场所氛围与情感价值", False),
    "公共艺术": ("场所氛围与情感价值", False),
    "社区归属感": ("场所氛围与情感价值", False),
    "多元文化融合度": ("场所氛围与情感价值", False),
}

# 负向指标列表
NEGATIVE_INDICATORS = [k for k, (_, neg) in AI_TO_FRAMEWORK.items() if neg]

print("=" * 60)
print("华安里感知模型 - BERT情感分类与6+1框架计算")
print("=" * 60)

# ============ 1. 加载数据 ============
print("\n[1/5] 加载数据...")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

records = []
headers = [cell.value for cell in ws[1]]

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3] is None:
        continue
    rec = {
        'id': str(row[0]) if row[0] is not None else '',
        'gps': str(row[1]) if row[1] is not None else '',
        'filename': str(row[2]) if row[2] is not None else '',
        'raw_text': str(row[3]),
    }
    # 解析 JSON
    text = rec['raw_text'].strip()
    if text.startswith('```json'):
        text = text[7:]
    if text.startswith('```'):
        text = text[3:]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()
    
    # 尝试修复常见JSON问题
    try:
        rec['analysis'] = json.loads(text)
        rec['parse_ok'] = True
    except json.JSONDecodeError:
        # 尝试补全截断的JSON
        try:
            # 找到最后一个完整的键值对
            fixed = text.rstrip().rstrip(',')
            # 尝试补全末尾的 }
            bracket_count = fixed.count('{') - fixed.count('}')
            fixed += '}' * bracket_count
            rec['analysis'] = json.loads(fixed)
            rec['parse_ok'] = True
            print(f"  修复成功: ID={rec['id']} ({rec['filename']})")
        except:
            rec['analysis'] = {}
            rec['parse_ok'] = False
            print(f"  解析失败: ID={rec['id']} ({rec['filename']})")
    
    records.append(rec)

print(f"  总记录数: {len(records)}")
print(f"  成功解析: {sum(1 for r in records if r['parse_ok'])}")
print(f"  解析失败: {sum(1 for r in records if not r['parse_ok'])}")

# ============ 2. 加载BERT模型 ============
print("\n[2/5] 加载BERT情感分析模型...")
device = 0 if torch.cuda.is_available() else -1
print(f"  设备: {'GPU' if device == 0 else 'CPU'}")

# 使用多语言BERT情感模型，输出1-5星评分
model_name = "nlptown/bert-base-multilingual-uncased-sentiment"
print(f"  模型: {model_name}")
sentiment_pipeline = pipeline(
    "sentiment-analysis",
    model=model_name,
    tokenizer=model_name,
    device=device,
    max_length=512,
    truncation=True,
)
print("  模型加载完成")

# ============ 3. BERT情感评分 ============
print("\n[3/5] BERT情感评分中...")
# 收集所有需要评分的文本
all_texts = []
text_index = []  # (record_idx, dim_name)

for i, rec in enumerate(records):
    if not rec['parse_ok']:
        continue
    for dim_name, dim_text in rec['analysis'].items():
        if dim_name in AI_TO_FRAMEWORK:
            all_texts.append(dim_text)
            text_index.append((i, dim_name))

print(f"  待评分文本数: {len(all_texts)} (={94}条×30维)")
print(f"  预计耗时: ~{len(all_texts) * 0.05:.0f}秒 (CPU)")

# 分批处理
BATCH_SIZE = 32
all_scores = []
start_time = time.time()
for batch_start in range(0, len(all_texts), BATCH_SIZE):
    batch_end = min(batch_start + BATCH_SIZE, len(all_texts))
    batch = all_texts[batch_start:batch_end]
    results = sentiment_pipeline(batch, batch_size=BATCH_SIZE)
    for r in results:
        # 将1-5星转换为0-1分数
        stars = int(r['label'][0])  # "4 stars" -> 4
        score = (stars - 1) / 4.0  # 1→0, 5→1
        all_scores.append(score)
    
    elapsed = time.time() - start_time
    done = batch_end
    total = len(all_texts)
    eta = (elapsed / done) * (total - done) if done > 0 else 0
    print(f"  进度: {done}/{total} ({100*done/total:.1f}%), 耗时: {elapsed:.0f}s, 预计剩余: {eta:.0f}s")

print(f"  评分完成! 总耗时: {time.time()-start_time:.0f}s")

# 将分数填入记录
for rec in records:
    rec['dim_scores'] = {}

# 对于负向指标，反转分数
for (rec_idx, dim_name), score in zip(text_index, all_scores):
    if dim_name in NEGATIVE_INDICATORS:
        score = 1.0 - score  # 反转：高分→低分（因为描述越负面，原始文本越详细，但我们要的是品质）
    records[rec_idx]['dim_scores'][dim_name] = score

# ============ 4. 映射到6+1框架并加权 ============
print("\n[4/5] 映射到6+1框架并计算加权得分...")

framework_dims = list(DIMENSION_WEIGHTS.keys())
for rec in records:
    rec['framework_scores'] = {}
    # 按框架维度聚合
    dim_accum = {d: [] for d in framework_dims}
    for ai_dim, score in rec.get('dim_scores', {}).items():
        if ai_dim in AI_TO_FRAMEWORK:
            fw_dim, _ = AI_TO_FRAMEWORK[ai_dim]
            dim_accum[fw_dim].append(score)
    
    # 计算每个框架维度的均值
    for d in framework_dims:
        scores = dim_accum[d]
        if scores:
            rec['framework_scores'][d] = sum(scores) / len(scores)
        else:
            rec['framework_scores'][d] = None
    
    # 维度7（针灸式更新潜力）：框架未定义AI直接映射指标
    # 临时方案：取前6维度的反值作为代理（感知越差→更新潜力越大）
    # 正式版本需结合TTI时空厚度指数计算
    if rec['framework_scores']['针灸式更新潜力'] is None:
        front6 = [rec['framework_scores'][d] for d in framework_dims[:6] if rec['framework_scores'][d] is not None]
        if front6:
            rec['framework_scores']['针灸式更新潜力'] = 1.0 - (sum(front6) / len(front6))
    
    # 计算综合得分（前6维度加权）
    weighted_sum = 0.0
    weight_sum = 0.0
    for d in framework_dims[:6]:  # 仅前6维度参与加权，+1维度独立评估
        s = rec['framework_scores'][d]
        w = DIMENSION_WEIGHTS[d]
        if s is not None:
            weighted_sum += s * w
            weight_sum += w
    
    rec['综合感知得分'] = weighted_sum / weight_sum if weight_sum > 0 else None
    # 重命名为6维度综合，+1独立
    rec['针灸式更新潜力'] = rec['framework_scores']['针灸式更新潜力']

# ============ 5. 输出结果 ============
print("\n[5/5] 输出结果...")

# 构建DataFrame
rows = []
for rec in records:
    row = {
        '编号': rec['id'],
        '文件名': rec['filename'],
        'GPS纬度': rec['gps'].split(',')[0] if ',' in rec['gps'] else '',
        'GPS经度': rec['gps'].split(',')[1] if ',' in rec['gps'] else '',
        '综合感知得分': rec['综合感知得分'],
    }
    for d in framework_dims:
        row[d] = rec['framework_scores'].get(d)
    rows.append(row)

df = pd.DataFrame(rows)

# 显示统计
print(f"\n  === 综合感知得分统计 ===")
valid_scores = [r['综合感知得分'] for r in records if r['综合感知得分'] is not None]
if valid_scores:
    print(f"  有效记录: {len(valid_scores)}")
    print(f"  均值: {sum(valid_scores)/len(valid_scores):.4f}")
    print(f"  最高: {max(valid_scores):.4f}")
    print(f"  最低: {min(valid_scores):.4f}")
    
    # 各维度统计
    print(f"\n  === 各维度得分统计 ===")
    for d in framework_dims:
        ds = [r['framework_scores'][d] for r in records if r['framework_scores'].get(d) is not None]
        if ds:
            print(f"  {d}: 均值={sum(ds)/len(ds):.4f}, 最高={max(ds):.4f}, 最低={min(ds):.4f}")

# 保存CSV
df.to_csv(OUTPUT_CSV, index=False, encoding='utf-8-sig')
print(f"\n  CSV已保存: {OUTPUT_CSV}")

# 保存Excel
df.to_excel(OUTPUT_EXCEL, index=False)
print(f"  Excel已保存: {OUTPUT_EXCEL}")

# 保存完整JSON（含每个子维度的得分）
json_output = []
for rec in records:
    jr = {
        'id': rec['id'],
        'filename': rec['filename'],
        'gps': rec['gps'],
        '综合感知得分': rec['综合感知得分'],
        'framework_scores': rec['framework_scores'],
        'ai_dim_scores': rec.get('dim_scores', {}),
    }
    json_output.append(jr)

with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(json_output, f, ensure_ascii=False, indent=2)
print(f"  JSON已保存: {OUTPUT_JSON}")

print(f"\n{'='*60}")
print("完成！")
print(f"{'='*60}")
